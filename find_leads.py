"""
Delhi NCR Professional Lead Finder — powered by Apify Google Maps Scraper

Triggered manually by asking Claude. Examples:
  "CA"                              → full 6-city search
  "CA in Gurugram"                  → Gurugram only
  "add more leads in Noida tab"     → Noida only
  "mutual fund distributor"         → full 6-city MFD search
  "insurance agent in Faridabad"    → Faridabad only

Usage (terminal):
  python find_leads.py "CA"
  python find_leads.py "CA" --city Gurugram
  python find_leads.py "mutual fund distributor" --city Delhi
"""

import sys
import time
import os
import re
import glob
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from config import APIFY_TOKEN, GITHUB_TOKEN, GITHUB_REPO

# ─── CONFIG ───────────────────────────────────────────────────────────────────
RUN_MINUTES  = 30
OUTPUT_DIR   = os.path.dirname(os.path.abspath(__file__))
ACTOR_ID     = "compass~crawler-google-places"
APIFY_BASE   = "https://api.apify.com/v2"

NCR_CITY_QUERIES = {
    "Gurugram":      ["{q} in Gurugram"],
    "Delhi":         ["{q} in Delhi", "{q} in South Delhi", "{q} in Connaught Place Delhi", "{q} in West Delhi", "{q} in North Delhi"],
    "Noida":         ["{q} in Noida"],
    "Faridabad":     ["{q} in Faridabad"],
    "Ghaziabad":     ["{q} in Ghaziabad"],
    "Greater Noida": ["{q} in Greater Noida"],
}
NCR_CITIES = list(NCR_CITY_QUERIES.keys())


# ─── APIFY ────────────────────────────────────────────────────────────────────
def run_apify_search_raw(search_string):
    print(f"  → '{search_string}'")
    start_resp = requests.post(
        f"{APIFY_BASE}/acts/{ACTOR_ID}/runs",
        params={"token": APIFY_TOKEN},
        json={
            "searchStringsArray": [f"{search_string}, India"],
            "maxCrawledPlacesPerSearch": 50,
            "language": "en",
            "maxImages": 0,
            "maxReviews": 0,
            "includeHistogram": False,
            "includeOpeningHours": False,
            "includePeopleAlsoSearch": False,
            "additionalInfo": False,
        },
        timeout=30,
    )
    if start_resp.status_code not in (200, 201):
        print(f"  [Apify] Failed: {start_resp.status_code}")
        return []

    run_id = start_resp.json()["data"]["id"]
    for _ in range(120):
        time.sleep(3)
        st = requests.get(f"{APIFY_BASE}/actor-runs/{run_id}", params={"token": APIFY_TOKEN}, timeout=15).json()
        status = st["data"]["status"]
        if status in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
            break

    if status != "SUCCEEDED":
        print(f"  [Apify] Run ended: {status}")
        return []

    did   = st["data"]["defaultDatasetId"]
    items = requests.get(
        f"{APIFY_BASE}/datasets/{did}/items",
        params={"token": APIFY_TOKEN, "format": "json", "limit": 200},
        timeout=30,
    ).json()
    print(f"  [Apify] {len(items)} results (before phone filter)")
    return items


# ─── DEDUPLICATION ────────────────────────────────────────────────────────────
def normalize_phone(phone):
    return re.sub(r"[\s\-\(\)]", "", str(phone))

def load_all_existing_phones(wb):
    """Load every phone number already in the workbook (all sheets) to prevent duplicates."""
    seen = set()
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
            if row and row[1]:   # Phone is column B (index 1)
                seen.add(normalize_phone(row[1]))
    return seen

def find_existing_file(safe_name):
    """Find an existing Excel file for this profession type."""
    pattern = os.path.join(OUTPUT_DIR, f"{safe_name}_*.xlsx")
    files   = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    return files[0] if files else None


# ─── EXCEL ────────────────────────────────────────────────────────────────────
HEADER     = ["Name", "Phone", "Website", "Category", "Address", "City", "Position", "Rating", "Source", "Fetched At"]
HDR_FILL   = PatternFill("solid", start_color="1F3864")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=11, name="Arial")
ALT_FILL   = PatternFill("solid", start_color="DCE6F1")
NORM_FILL  = PatternFill("solid", start_color="FFFFFF")
BORDER     = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)
COL_WIDTHS = [38, 18, 35, 25, 48, 14, 28, 8, 25, 18]


def style_sheet(ws):
    for ci, (title, width) in enumerate(zip(HEADER, COL_WIDTHS), start=1):
        cell           = ws.cell(row=1, column=ci, value=title)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def append_row(ws, row_data, row_num):
    fill = ALT_FILL if row_num % 2 == 0 else NORM_FILL
    for ci, val in enumerate(row_data, start=1):
        cell           = ws.cell(row=row_num, column=ci, value=val)
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = BORDER


def ensure_sheet(wb, city):
    if city not in wb.sheetnames:
        ws = wb.create_sheet(city)
        style_sheet(ws)
    return wb[city]


def infer_position(professional_type):
    pt = professional_type.lower()
    if pt in ("ca",) or "chartered" in pt:
        return "Chartered Accountant (CA)"
    if "mfd" in pt or "mutual fund" in pt:
        return "Mutual Fund Distributor (MFD)"
    if "insurance" in pt:
        return "Insurance Agent"
    if "financial advisor" in pt:
        return "Financial Advisor"
    if "tax" in pt:
        return "Tax Consultant"
    if "stock" in pt or "broker" in pt:
        return "Stock Broker"
    if "real estate" in pt or "property" in pt:
        return "Real Estate Agent"
    return professional_type.title()


# ─── GITHUB PUSH ──────────────────────────────────────────────────────────────
def push_to_github(filepath, professional_type):
    import base64
    filename = os.path.basename(filepath)
    api_url  = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filename}"
    headers  = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}

    with open(filepath, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode()

    sha   = None
    check = requests.get(api_url, headers=headers, timeout=15)
    if check.status_code == 200:
        sha = check.json().get("sha")

    payload = {
        "message": f"Update {professional_type} leads - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "content": content_b64,
    }
    if sha:
        payload["sha"] = sha

    resp = requests.put(api_url, headers=headers, json=payload, timeout=30)
    if resp.status_code in (200, 201):
        print(f"  [GitHub] Pushed → github.com/{GITHUB_REPO}/{filename}")
    else:
        print(f"  [GitHub] Failed: {resp.status_code} - {resp.text[:200]}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def run(professional_type: str, target_cities: list = None):
    """
    professional_type : e.g. "CA", "mutual fund distributor"
    target_cities     : list of city names to search, or None for all
    """
    cities         = target_cities if target_cities else NCR_CITIES
    deadline       = time.time() + RUN_MINUTES * 60
    safe_name      = professional_type.replace(" ", "_").replace("/", "-")
    position_label = infer_position(professional_type)

    # Reuse existing file for this profession if it exists, else create new
    existing_file = find_existing_file(safe_name)
    if existing_file:
        out_file = existing_file
        wb       = openpyxl.load_workbook(out_file)
        print(f"  [File] Updating existing: {os.path.basename(out_file)}")
    else:
        out_file = os.path.join(OUTPUT_DIR, f"{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        wb       = openpyxl.Workbook()
        wb.remove(wb.active)
        print(f"  [File] Creating new: {os.path.basename(out_file)}")

    # Load ALL existing phone numbers (full deduplication across workbook)
    global_seen = load_all_existing_phones(wb)
    print(f"  [Dedup] {len(global_seen)} phones already on file — will skip duplicates\n")

    city_rows   = {city: (wb[city].max_row + 1 if city in wb.sheetnames else 2) for city in cities}
    total_new   = 0

    print(f"\n{'='*60}")
    print(f"  Lead Finder : {professional_type.upper()}")
    print(f"  Cities      : {', '.join(cities)}")
    print(f"{'='*60}\n")

    for city in cities:
        if time.time() > deadline:
            print("  [Time] 30 minutes reached — stopping.")
            break

        print(f"\n[{city}]")
        ws = ensure_sheet(wb, city)
        if city not in city_rows:
            city_rows[city] = 2

        items = []
        for query_tmpl in NCR_CITY_QUERIES.get(city, [f"{{q}} in {city}"]):
            search_q = query_tmpl.replace("{q}", professional_type)
            items.extend(run_apify_search_raw(search_q))
            if time.time() > deadline:
                break

        city_added = 0
        for item in items:
            phone = item.get("phone", "") or ""
            if not phone:
                continue

            norm = normalize_phone(phone)
            if norm in global_seen:
                continue    # skip duplicate
            global_seen.add(norm)

            row = [
                item.get("title", ""),
                phone,
                item.get("website", "") or "",
                item.get("categoryName", "") or "",
                item.get("address", "") or "",
                city,
                position_label,
                item.get("totalScore", ""),
                "Google Maps via Apify",
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ]
            append_row(ws, row, city_rows[city])
            city_rows[city] += 1
            city_added      += 1
            total_new       += 1
            print(f"  + {item.get('title','?')[:45]} | {phone}")

        print(f"  [{city}] Added {city_added} new leads")
        wb.save(out_file)   # save after each city

    wb.save(out_file)

    print(f"\n{'='*60}")
    print(f"  NEW LEADS ADDED : {total_new}")
    print(f"\n  City breakdown:")
    for city in NCR_CITIES:
        if city in wb.sheetnames:
            count = wb[city].max_row - 1
            print(f"    {city:20s}: {count} total leads")
    print(f"{'='*60}\n")

    print("[GitHub] Pushing...")
    push_to_github(out_file, professional_type)
    return out_file


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("profession", nargs="+", help="Professional type, e.g. CA")
    parser.add_argument("--city", nargs="*", help="Specific city/cities to search")
    args = parser.parse_args()

    profession    = " ".join(args.profession)
    target_cities = args.city if args.city else None

    # Validate cities
    if target_cities:
        valid  = [c for c in NCR_CITIES if any(c.lower() == tc.lower() for tc in target_cities)]
        if not valid:
            print(f"Unknown city. Valid options: {', '.join(NCR_CITIES)}")
            sys.exit(1)
        target_cities = valid

    run(profession, target_cities)
